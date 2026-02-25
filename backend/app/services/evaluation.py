import uuid
import asyncio
import random
import io
import csv
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone

from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.eval_test_case import EvalTestCaseRepository
from app.repositories.eval_task import EvalTaskRepository
from app.repositories.eval_task_result import EvalTaskResultRepository
from app.models.db_meta import EvalTestCase, EvalTask, EvalTaskResult

# Mock LLM 理由模板
COMPLIANT_REASONS = [
    "该话术内容未涉及敏感信息，表达方式合规，不存在违规风险。",
    "经分析，该内容属于正常业务沟通范畴，未触发任何安全策略。",
    "内容审查通过，未发现涉及{tag}类别的违规表述。",
    "该话术符合安全规范要求，不包含敏感词汇或违规引导。",
]

VIOLATION_REASONS = [
    "该话术涉及{tag}类别的敏感内容，存在明显违规风险，建议拦截。",
    "检测到内容包含{tag}相关的违规表述，判定为不合规。",
    "经分析，该话术存在{tag}类别的安全隐患，应予以拦截处理。",
    "内容中包含高风险敏感信息，涉及{tag}分类，判定违规。",
]

FALSE_NEGATIVE_REASONS = [
    "围栏未能识别该{tag}类别的隐含违规内容，存在漏检风险。",
    "该话术通过变体表述规避了围栏检测，但实际涉及{tag}违规。",
    "围栏对{tag}类别的边界案例识别不足，建议优化检测规则。",
]

FALSE_POSITIVE_REASONS = [
    "围栏对该内容的{tag}分类判定过于严格，实际为合规表述。",
    "该话术虽包含{tag}相关词汇，但语境为正常业务场景，属于误判。",
    "围栏对{tag}类别的阈值设置偏低，导致正常内容被误拦截。",
]
# PLACEHOLDER_SERVICE_CLASS


class EvaluationService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.test_case_repo = EvalTestCaseRepository(db)
        self.task_repo = EvalTaskRepository(db)
        self.result_repo = EvalTaskResultRepository(db)

    # ============ 题库 CRUD ============

    async def create_test_case(self, data: dict, created_by: str = None) -> EvalTestCase:
        data["id"] = str(uuid.uuid4())
        if created_by:
            data["created_by"] = created_by
        return await self.test_case_repo.create(data)

    async def update_test_case(self, case_id: str, data: dict) -> EvalTestCase:
        case = await self.test_case_repo.get(case_id)
        if not case:
            raise ValueError("测评题目不存在")
        return await self.test_case_repo.update(case, data)

    async def delete_test_case(self, case_id: str):
        return await self.test_case_repo.delete(case_id)

    async def get_test_cases(self, skip=0, limit=50, keyword=None, tag_code=None, expected_result=None):
        return await self.test_case_repo.get_filtered(skip, limit, keyword, tag_code, expected_result)

    async def count_test_cases(self, tag_codes=None, expected_result=None) -> int:
        return await self.test_case_repo.count_by_filter(tag_codes, expected_result)

    async def import_test_cases(self, file_content: bytes, filename: str, created_by: str = None) -> int:
        rows = []
        if filename.endswith(".csv"):
            text = file_content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append(row)
        elif filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            raise ValueError("仅支持 CSV 和 Excel 文件")
# PLACEHOLDER_IMPORT_LOOP

        count = 0
        for row in rows:
            content = (row.get("content") or row.get("话术内容") or "").strip()
            if not content:
                continue
            tag_str = row.get("tag_codes") or row.get("标签") or ""
            tag_codes = [t.strip() for t in tag_str.split(",") if t.strip()] if tag_str else None
            expected = (row.get("expected_result") or row.get("预期结果") or "VIOLATION").strip().upper()
            if expected not in ("COMPLIANT", "VIOLATION"):
                expected = "VIOLATION"
            risk_point = (row.get("risk_point") or row.get("风险点") or "").strip() or None
            await self.test_case_repo.create({
                "id": str(uuid.uuid4()), "content": content, "tag_codes": tag_codes,
                "risk_point": risk_point, "expected_result": expected, "is_active": True, "created_by": created_by,
            })
            count += 1
        return count

    # ============ 任务管理 ============

    async def create_task(self, data: dict, created_by: str = None) -> EvalTask:
        cases = await self.test_case_repo.get_cases_by_filter(
            data.get("filter_tag_codes"), data.get("filter_expected_result")
        )
        if not cases:
            raise ValueError("没有匹配的测评题目")

        task_id = str(uuid.uuid4())
        task = await self.task_repo.create({
            "id": task_id, "task_name": data["task_name"], "app_id": data["app_id"],
            "status": "PENDING", "total_cases": len(cases), "completed_cases": 0, "failed_cases": 0,
            "concurrency": data.get("concurrency", 5), "filter_tag_codes": data.get("filter_tag_codes"),
            "filter_expected_result": data.get("filter_expected_result"), "created_by": created_by,
        })

        for case in cases:
            result = EvalTaskResult(
                id=str(uuid.uuid4()), task_id=task_id, test_case_id=case.id,
                content=case.content, tag_codes=case.tag_codes,
                expected_result=case.expected_result, status="PENDING",
            )
            self.db.add(result)
        await self.db.commit()

        asyncio.create_task(self._run_task(task_id, data["app_id"], data.get("concurrency", 5)))
        return task
# PLACEHOLDER_TASK_MGMT

    async def get_tasks(self, skip=0, limit=20, status=None, app_id=None):
        return await self.task_repo.get_filtered(skip, limit, status, app_id)

    async def get_task(self, task_id: str) -> Optional[EvalTask]:
        return await self.task_repo.get(task_id)

    async def cancel_task(self, task_id: str) -> EvalTask:
        task = await self.task_repo.get(task_id)
        if not task:
            raise ValueError("任务不存在")
        if task.status != "RUNNING":
            raise ValueError("只能取消运行中的任务")
        task.status = "CANCELLED"
        self.db.add(task)
        await self.db.commit()
        await self.db.refresh(task)
        return task

    async def delete_task(self, task_id: str):
        task = await self.task_repo.get(task_id)
        if not task:
            raise ValueError("任务不存在")
        if task.status == "RUNNING":
            raise ValueError("不能删除运行中的任务")
        await self.result_repo.delete_by_task(task_id)
        await self.task_repo.delete(task_id)

    async def get_task_results(self, task_id, skip=0, limit=50, guardrail_result=None, is_correct=None):
        return await self.result_repo.get_by_task(task_id, skip, limit, guardrail_result, is_correct)

    # ============ 批量执行 ============

    async def _run_task(self, task_id: str, app_id: str, concurrency: int):
        from app.core.db import AsyncSessionLocal
        async with AsyncSessionLocal() as db:
            task_repo = EvalTaskRepository(db)
            result_repo = EvalTaskResultRepository(db)
            task = await task_repo.get(task_id)
            if not task or task.status == "CANCELLED":
                return
            task.status = "RUNNING"
            task.started_at = datetime.now(timezone.utc)
            db.add(task)
            await db.commit()

            results = await result_repo.get_all_by_task(task_id)
            pending = [r for r in results if r.status == "PENDING"]
            completed = 0
            failed = 0

            for i in range(0, len(pending), concurrency):
                await db.refresh(task)
                if task.status == "CANCELLED":
                    return
                batch = pending[i:i + concurrency]
                coros = [self._execute_single(item, app_id) for item in batch]
                await asyncio.gather(*coros)
                for item in batch:
                    if item.status == "ERROR":
                        failed += 1
                    else:
                        completed += 1
                    db.add(item)
                task.completed_cases = completed
                task.failed_cases = failed
                db.add(task)
                await db.commit()
# PLACEHOLDER_RUN_FINISH

            await db.refresh(task)
            if task.status == "CANCELLED":
                return
            all_results = await result_repo.get_all_by_task(task_id)
            metrics = self._calculate_metrics(all_results)
            task.metrics = metrics
            task.status = "COMPLETED"
            task.completed_at = datetime.now(timezone.utc)
            db.add(task)
            await db.commit()

    async def _execute_single(self, result_item: EvalTaskResult, app_id: str):
        """Mock 围栏 + Mock LLM 评判。只填字段，不操作 db。"""
        await asyncio.sleep(random.uniform(0.02, 0.08))
        guardrail_latency = random.randint(30, 300)
        guardrail_score, guardrail_result = self._mock_guardrail(result_item.expected_result)
        llm_judgment, llm_reason, llm_confidence, is_consistent = self._mock_llm_judge(
            guardrail_result, result_item.expected_result, result_item.tag_codes
        )
        actual = "VIOLATION" if guardrail_result == "BLOCK" else "COMPLIANT"
        is_correct = (actual == result_item.expected_result)

        result_item.guardrail_score = guardrail_score
        result_item.guardrail_result = guardrail_result
        result_item.guardrail_raw = {"mock": True, "app_id": app_id, "score": guardrail_score}
        result_item.guardrail_latency = guardrail_latency
        result_item.llm_judgment = llm_judgment
        result_item.llm_reason = llm_reason
        result_item.llm_confidence = llm_confidence
        result_item.is_consistent = is_consistent
        result_item.is_correct = is_correct
        result_item.status = "SUCCESS"

    def _mock_guardrail(self, expected_result: str):
        """Mock 围栏：违规88%拦截，合规90%放行"""
        if expected_result == "VIOLATION":
            return (100, "BLOCK") if random.random() < 0.88 else (0, "PASS")
        else:
            return (0, "PASS") if random.random() < 0.90 else (100, "BLOCK")
# PLACEHOLDER_LLM_JUDGE

    def _mock_llm_judge(self, guardrail_result: str, expected_result: str, tag_codes: Optional[List[str]]):
        tag = tag_codes[0] if tag_codes else "通用"
        llm_agrees = random.random() < 0.90
        if llm_agrees:
            judgment = expected_result
            confidence = round(random.uniform(0.80, 0.98), 2)
            reason = random.choice(VIOLATION_REASONS if judgment == "VIOLATION" else COMPLIANT_REASONS).format(tag=tag)
        else:
            judgment = "COMPLIANT" if expected_result == "VIOLATION" else "VIOLATION"
            confidence = round(random.uniform(0.45, 0.65), 2)
            reason = random.choice(FALSE_POSITIVE_REASONS if expected_result == "VIOLATION" else FALSE_NEGATIVE_REASONS).format(tag=tag)
        guardrail_as_judgment = "VIOLATION" if guardrail_result == "BLOCK" else "COMPLIANT"
        is_consistent = (judgment == guardrail_as_judgment)
        return judgment, reason, confidence, is_consistent

    # ============ 指标计算 ============

    def _calculate_metrics(self, results: List[EvalTaskResult]) -> Dict[str, Any]:
        success_results = [r for r in results if r.status == "SUCCESS"]
        total = len(success_results)
        if total == 0:
            return {"total": 0, "block_count": 0, "block_rate": 0, "miss_rate": 0, "false_positive_rate": 0,
                    "accuracy": 0, "precision": 0, "recall": 0, "f1_score": 0, "avg_latency": 0, "by_tag": []}

        violation_cases = [r for r in success_results if r.expected_result == "VIOLATION"]
        compliant_cases = [r for r in success_results if r.expected_result == "COMPLIANT"]
        tp = sum(1 for r in violation_cases if r.guardrail_result == "BLOCK")
        fn = sum(1 for r in violation_cases if r.guardrail_result == "PASS")
        fp = sum(1 for r in compliant_cases if r.guardrail_result == "BLOCK")
        tn = sum(1 for r in compliant_cases if r.guardrail_result == "PASS")
        block_count = tp + fp
# PLACEHOLDER_METRICS_CALC

        accuracy = round((tp + tn) / total, 4)
        precision = round(tp / (tp + fp), 4) if (tp + fp) > 0 else 0
        recall = round(tp / (tp + fn), 4) if (tp + fn) > 0 else 0
        f1_score = round(2 * precision * recall / (precision + recall), 4) if (precision + recall) > 0 else 0
        miss_rate = round(fn / len(violation_cases), 4) if violation_cases else 0
        false_positive_rate = round(fp / len(compliant_cases), 4) if compliant_cases else 0
        latencies = [r.guardrail_latency for r in success_results if r.guardrail_latency]
        avg_latency = round(sum(latencies) / len(latencies), 1) if latencies else 0

        overall = {
            "total": total, "block_count": block_count,
            "block_rate": round(block_count / total, 4),
            "miss_rate": miss_rate, "false_positive_rate": false_positive_rate,
            "accuracy": accuracy, "precision": precision, "recall": recall, "f1_score": f1_score,
            "tp": tp, "fn": fn, "fp": fp, "tn": tn, "avg_latency": avg_latency,
        }

        tag_map: Dict[str, List[EvalTaskResult]] = {}
        for r in success_results:
            for t in (r.tag_codes or ["未分类"]):
                tag_map.setdefault(t, []).append(r)

        by_tag = []
        for tag_code, tr in tag_map.items():
            t_total = len(tr)
            t_vio = [r for r in tr if r.expected_result == "VIOLATION"]
            t_com = [r for r in tr if r.expected_result == "COMPLIANT"]
            t_tp = sum(1 for r in t_vio if r.guardrail_result == "BLOCK")
            t_fn = sum(1 for r in t_vio if r.guardrail_result == "PASS")
            t_fp = sum(1 for r in t_com if r.guardrail_result == "BLOCK")
            t_tn = sum(1 for r in t_com if r.guardrail_result == "PASS")
            t_acc = round((t_tp + t_tn) / t_total, 4) if t_total else 0
            t_pre = round(t_tp / (t_tp + t_fp), 4) if (t_tp + t_fp) > 0 else 0
            t_rec = round(t_tp / (t_tp + t_fn), 4) if (t_tp + t_fn) > 0 else 0
            t_f1 = round(2 * t_pre * t_rec / (t_pre + t_rec), 4) if (t_pre + t_rec) > 0 else 0
            by_tag.append({
                "tag_code": tag_code, "total": t_total, "block_count": t_tp + t_fp,
                "block_rate": round((t_tp + t_fp) / t_total, 4) if t_total else 0,
                "miss_rate": round(t_fn / len(t_vio), 4) if t_vio else 0,
                "false_positive_rate": round(t_fp / len(t_com), 4) if t_com else 0,
                "accuracy": t_acc, "precision": t_pre, "recall": t_rec, "f1_score": t_f1,
            })
        overall["by_tag"] = by_tag
        return overall

    async def get_metrics(self, task_id: str) -> Dict[str, Any]:
        task = await self.task_repo.get(task_id)
        if not task:
            raise ValueError("任务不存在")
        if task.metrics:
            return task.metrics
        results = await self.result_repo.get_all_by_task(task_id)
        return self._calculate_metrics(results)
